import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import random

# Create a list of electronic products
products = [
    {"id": 1, "name": "Laptop", "price": 1000},
    {"id": 2, "name": "Smartphone", "price": 500},
    {"id": 3, "name": "Tablet", "price": 300},
    {"id": 4, "name": "Headphones", "price": 100},
    {"id": 5, "name": "Smartwatch", "price": 200},
]

# Create a new workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set the headers
headers = ["ID", "Name", "Quantity", "Price"]
for col_num, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col_num, value=header)

# Generate sales data for 100 rows
for row_num in range(2, 102):
    product = random.choice(products)
    quantity = random.randint(1, 10)
    price = product["price"]
    total_price = quantity * price

    sheet.cell(row=row_num, column=1, value=product["id"])
    sheet.cell(row=row_num, column=2, value=product["name"])
    sheet.cell(row=row_num, column=3, value=quantity)
    sheet.cell(row=row_num, column=4, value=total_price)

# Save the workbook to 'c:\work\sales.xlsx'
workbook.save('c:\\work\\sales.xlsx')

print("Sales list has been created and saved as 'c:\\work\\sales.xlsx'.")
